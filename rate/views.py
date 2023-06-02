import csv
from django.shortcuts import render, redirect
from django.urls import reverse
from django.http import HttpResponse, HttpResponseRedirect
from django.http import JsonResponse
from io import StringIO
from django.conf import settings
from django.core.files.storage import FileSystemStorage
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Protection, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import re
from datetime import date, datetime
from dateutil import parser
from django.db.models import Q
from django.contrib.auth import views as auth_views
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.db import connection, transaction
from zipfile import ZipFile, ZIP_DEFLATED
from .models import *
from .utility import *
from .services import *
from .log import error, info

# Create your views here.
@login_required
def home(request):
    # Save user id in session
    request.session["RATE_USER_ID"] = request.user.id
    
    info("========In home========")
    info("Logged in user name " + request.user.username)

    # If staff user then redirect to admin customer page
    if request.user.is_staff:
        request.session["customer_key"] = ""
        info("Logged in user is a staff user")
        return redirect("master-customers") 
 
    # User is a customer
    # Get the profile for the logged in user
    profile = Profile()
    try:        
        profile = request.user.profile
    except Exception as e:
        # If profile doesn't exist show error
        error("home - Unable to fetch profile " + str(e))
        return render(request, "rate/common/error_page.html", {"error_message" : "Logged-in user is not associated to any customer! Please contact the administrator."})
    
    # Store customer key, customer name and domain from user profile, and store them in session
    info("Customer name " + profile.customer.customer_name)
    request.session["customer_key"] = profile.customer.customer_key
    request.session["customer_name"] = profile.customer.customer_name
    request.session["domain"] = profile.customer.domain

    # Check the previus URL    
    # info("Previous URL " + str(request.META.get('HTTP_REFERER')))
    # If the user logged in from the generic login page (i.e. not customer specific URL), then show error and provide the correct login URL for the logged-in user
    # if not profile.customer.customer_key in str(request.META.get('HTTP_REFERER')):
    #     error("home - User logged in from generic login page")
    #     return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to access this content. Please use following URL to login.", "url" : request.build_absolute_uri() + profile.customer.customer_key + "/login/", "show_home_link" : "N"})    

    if "logged_in_from_url" in request.session:
        # Check the previus URL    
        info("logged_in_from_url " + str(request.session["logged_in_from_url"]))
        # If the user logged in from the generic login page (i.e. not customer specific URL), then show error and provide the correct login URL for the logged-in user
        if not profile.customer.customer_key in str(request.session["logged_in_from_url"]):
            error("home - User logged in from generic login page")
            return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to access this content. Please use following URL to login.", "url" : request.build_absolute_uri() + profile.customer.customer_key + "/login/", "show_home_link" : "N"})    
    else:
        info("logged_in_from_url doesn't exist in session, redirect to login")
        return HttpResponseRedirect("login")

    # Check if the customer attached to the user/profile is active
    customer = Customer()
    try:
        today = date.today()
        customer = Customer.objects.get((Q(end_date__gt=today) | Q(end_date__isnull=True)) & Q(customer_key=profile.customer.customer_key))
    except:
        # Show error if the customer is end-dated
        return render(request, "rate/common/error_page.html", {"error_message" : "Logged-in user is not associated to any ACTIVE customer! Please contact the administrator."})

    # If the logged-in user is customer-admin, then show manager home page
    # Else, show user home page
    if profile.user_type == "USER":
        info("User type is USER, navigating to select-instance page")
        return redirect("select-instance", profile.customer.customer_key)
    else:
        info("User type is ADMIN, navigating to manage-dashboard page")
        # Check if ADMIN user of the customer is supposed to have access of rate upload screens, save that in session
        request.session["allow_admin_rate_upload"] = customer.allow_admin_rate_upload
        return redirect("manager-dashboard", profile.customer.customer_key)


@login_required
def select_instance(request, key):
    info("========In select_instance========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    # If the user is not a staff user, and if the user is customer ADMIN but is not suppsed to have access of rate upload screens, then show error
    if request.user.is_staff == False:
        if request.user.profile.user_type == "ADMIN" and request.session["allow_admin_rate_upload"] == "N":
            return render(request, "rate/common/error_page.html", {"error_message" : "ADMIN user is not authorized to view this page."})

    # Get customer object
    customer_key = ""
    if "customer_key" in request.session:
        customer_key = request.session["customer_key"]
        info("customer key from session " + customer_key)

    customer = Customer.objects.none
    try:
        customer = Customer.objects.get(customer_key=customer_key)
    except:
        return render(request, "rate/common/error_page.html", {"error_message" : "Logged-in user is not associated to any ACTIVE customer! Please contact the administrator."}) 

    if request.method == "POST":
        # Get the instance that user selected and store it in session
        instance_id = int(request.POST.get("instance_id", ""))
        request.session["instance_id"] = instance_id
        instance = Instance.objects.get(id=instance_id)
        request.session["instance_name"] = str(instance.instance_name)
        # Get the template batch that user selected and store it in session
        template_batch_id = request.POST.get("template_batch", "")
        template_id = -1
        if template_batch_id != "": # User selected a template batch
            request.session["template_batch_id"] = template_batch_id
            # Get template batch
            batch = Batch.objects.get(id=template_batch_id, customer=customer)
            # Get the template and store it in session
            template_id = batch.template.id
            request.session["template_id"] = str(template_id)
        else: # User selected offering type and rate type
            # Get the template that user selected and store it in session
            template_id = int(request.POST.get("rate_type", ""))
            request.session["template_id"] = str(template_id)
        
        # Get the details of the selected template
        template = Template.objects.get(id=template_id)
        # Store offering type and rate type in session
        request.session["offering_type_name"] = template.offering_type_name
        request.session["offering_type_description"] = template.offering_type_description
        request.session["rate_type_name"] = template.rate_type_name
        if template.group_name: 
            request.session["group_name"] = template.group_name

        info("instance_id " + str(instance_id))
        info("instance_name " + instance.instance_name)
        info("template_id " + str(template_id))
        info("offering_type_name " + template.offering_type_name)
        info("offering_type_description " + template.offering_type_description)
        info("group_name " + template.group_name if template.group_name else "")
        info("rate_type_name " + template.rate_type_name)
        info("template_batch_id " + template_batch_id)  
        
        # Get the group associated to the template
        group = Group.objects.get(customer=customer, group_name=template.group_name)
        # Get all attributes for that group and store them in session
        request.session["attribute1"] = group.attribute1
        request.session["attribute2"] = group.attribute2
        request.session["attribute3"] = group.attribute3
        request.session["attribute4"] = group.attribute4
        request.session["attribute5"] = group.attribute5
        request.session["attribute6"] = group.attribute6
        request.session["attribute7"] = group.attribute7
        request.session["attribute8"] = group.attribute8
        request.session["attribute9"] = group.attribute9
        request.session["attribute10"] = group.attribute10

        # If Create Template button was pressed then show the options for downloading the template
        if "CreateTemplate" == request.POST.get("action", ""):
            # Navigate to generate-template page
            return redirect("generate-template-generic", request.session["customer_key"])

        # If Upload Template button was pressed then show the file upload option
        else:
            # Navigate to upload-template page
            return redirect("upload-template", request.session["customer_key"]) 
    else:
        # Get customer key from session
        customer_key = ""
        if "customer_key" in request.session:
            customer_key = request.session["customer_key"]

        today = date.today()

        # Validate if the customer is active
        customer = Customer()
        try:
            customer = Customer.objects.get((Q(end_date__gt=today) | Q(end_date__isnull=True)) & Q(customer_key=customer_key))
        except:
            return render(request, "rate/common/error_page.html", {"error_message" : "Logged-in user is not associated to any ACTIVE customer! Please contact the administrator."})
        
        # Fetch the list of all active instances granted to the logged-in user for the current customer
        accesses = UserAccess.objects.filter((Q(end_date__gt=today) | Q(end_date__isnull=True)) & Q(user=request.user) & Q(customer=customer))

        if request.user.is_staff == True:
            accesses = UserAccess.objects.filter((Q(end_date__gt=today) | Q(end_date__isnull=True)) & Q(customer=customer))
        
        # Show error if the user doesn't have access to any instance
        if accesses.count() == 0:
            return render(request, "rate/common/error_page.html", {"error_message" : "Logged-in user doesn't have access to any OTM instance! Please contact the administrator."})

        # Get the list of instance ids where the user has access
        instance_id_list = [(access.instance.id) for access in accesses]

        # Fetch the list of all active instances for the current customer and user
        instances = Instance.objects.filter((Q(end_date__gt=today) | Q(end_date__isnull=True)) & Q(customer=customer) & Q(id__in=instance_id_list)).order_by("instance_name")
        
        if instances.count() == 0:
            return render(request, "rate/common/error_page.html", {"error_message" : "No active instance found! Please contact the administrator."})

        # Get the statuses of all instances
        instances_ids = [(instance.id) for instance in instances]
        instances_names = [(instance.instance_name) for instance in instances]
        instances_statues = [(get_instance_status(instance)) for instance in instances]
        instances_with_statuses = []
        for i, instancesId in enumerate(instances_ids):
            instances_with_statuses.append({"id" : instancesId, "instance_name" : instances_names[i], "status" : instances_statues[i]})
            
        info(str(instances_with_statuses))

        group_names = []
        # If the user is a staff user, then get all groups for the customer, else get only the groups attached to the user
        if request.user.is_staff == True:
            groups = Group.objects.filter(Q(customer=customer))
            group_names = [(group.group_name) for group in groups]
        else:
            group_names = request.user.profile.group_name.upper().split(",")
            group_names = [(group_name.strip()) for group_name in group_names]
            groups = Group.objects.filter(Q(group_name__in=group_names) & Q(customer=customer))
            group_names = [(group.group_name) for group in groups]

        info(str(group_names))

        # Get all active templates for the customer
        offerings = []
        if request.user.is_staff == True:
            # If the user is a staff user, then get all templates for the customer
            offerings = Template.objects.all().filter((Q(end_date__gt=today) | Q(end_date__isnull=True)) & Q(customer=customer) & Q(enabled="Y")).values("offering_type_name", "offering_type_description", "group_name").distinct().order_by("offering_type_name", "group_name")
        else:
            if request.user.profile.user_type == "ADMIN":
                # If the user is customer ADMIN user, then get all templates for the customer
                offerings = Template.objects.all().filter((Q(end_date__gt=today) | Q(end_date__isnull=True)) & Q(customer=customer) & Q(enabled="Y")).values("offering_type_name", "offering_type_description", "group_name").distinct().order_by("offering_type_name", "group_name")
            else:
                # If the user is customer user, then get the templates for the user's groups
                offerings = Template.objects.all().filter((Q(end_date__gt=today) | Q(end_date__isnull=True)) & Q(customer=customer) & Q(enabled="Y") & (Q(group_name__in=group_names))).values("offering_type_name", "offering_type_description", "group_name").distinct().order_by("offering_type_name", "group_name")
        
        # Sort the templates so that the ones with DEFAULT group come first
        offerings = sorted(offerings, key=lambda ofr: "" if ofr["group_name"] == "DEFAULT" else ofr["group_name"])
        info(str(offerings))

        # If no active rate type/template exists then show error
        if len(offerings) == 0:
            return render(request, "rate/common/error_page.html", {"error_message" : "No offering types has been defined! Please contact the administrator."})
        
        # Get template batches
        batches = Batch.objects.none()
        if customer.view_all_template_batches == "Y":
            batches = Batch.objects.filter(customer=customer).order_by('-id')
        else:
            batches = Batch.objects.filter(customer=customer, created_by=request.user.username).order_by('-id')
        info("Total template batches found " + str(len(batches)))
        return render(request, "rate/customer/user/select_instance_page.html", {"customer" : customer, "instances" : instances_with_statuses, "offerings" : offerings, "batches" : batches})


@login_required
def generate_template_generic(request, key):
    info("========In generate_template_generic========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    # If the user is not a staff user, and if the user is customer ADMIN but is not suppsed to have access of rate upload screens, then show error
    if request.user.is_staff == False:
        if request.user.profile.user_type == "ADMIN" and request.session["allow_admin_rate_upload"] == "N":
            return render(request, "rate/common/error_page.html", {"error_message" : "ADMIN user is not authorized to view this page."})

    if request.method == "POST": 
        customer_key = ""
        template_id = 0
        offering_type_name = ""
        offering_type_description = ""
        rate_type_name = ""

        # Get values from session
        if "customer_key" in request.session:
            customer_key = request.session["customer_key"]
        if "template_id" in request.session:
            template_id = int(request.session["template_id"])
        if "offering_type_name" in request.session:
            offering_type_name = request.session["offering_type_name"]
        if "offering_type_description" in request.session:
            offering_type_description = request.session["offering_type_description"]
        if "rate_type_name" in request.session:
            rate_type_name = request.session["rate_type_name"]

        info("customer_key " + customer_key)
        info("template_id " + str(template_id))
        info("offering_type_name " + offering_type_name)
        info("offering_type_description " + offering_type_description)
        info("rate_type_name " + rate_type_name)

        # Get customer and rate type objects
        try:
            customer = Customer.objects.get(customer_key=customer_key)
        except:
            return render(request, "rate/common/error_page.html", {"error_message" : "Logged-in user is not associated to any customer! Please contact the administrator."})
            
        # Get instance details
        instance_id = 0
        domain = ""
        if "instance_id" in request.session:
            instance_id = int(request.session["instance_id"])
        if "domain" in request.session:
            domain = request.session["domain"]

        info("instance_id " + str(instance_id))
        info("domain " + domain)

        # Get the instance details of the selected instance
        instance = Instance.objects.get(id=instance_id, customer=customer)
        
        # Get the template details of the selected template
        try:
            template = Template.objects.get(id=template_id)
        except:
            return render(request, "rate/common/error_page.html", {"error_message" : "No rate found for the customer! Please contact the administrator."})
            
        # Get list of static columns
        template_static_columns = TemplateStaticColumn.objects.filter(customer=customer, template=template, enabled="Y", include_in_template="Y").order_by("position")
        if template_static_columns.count() == 0:
            return render(request, "rate/common/error_page.html", {"error_message" : "No static columns found for the template! Please contact the administrator."})

        # Get how many records were populated in the table
        table_record_count = int(request.POST.get("table_record_count", "")) - 1 # It includes table header row, so less it by one
        info("Total Rows Filled up " + str(table_record_count))
        
        # Create new Excel workbook and sheet
        wb = Workbook()  
        sheet = wb.active
        data_starts_from_row = 4 # 1st row contains sequence, 2nd row contains master column name, 3rd row contains template column name
        concatenated_charges = ""

        # Loop through each row that user filled up in the page
        for rowIndex in range(0, table_record_count):
            record_count = int(request.POST.get("excel_row_count_" + str(rowIndex), "")) # Get the count of rows that user wants to be populated in the Excel
            info("Count of Excel rows " + str(record_count))

            selected_charges = request.POST.getlist("charges_0")  # index 0 has been used to ignore selected charges from other rows
            info("Selected charges " + str(selected_charges))
            concatenated_charges = ""

            weightbreak_profile_name = request.POST.get("weight_break_profile_0", "")  # index 0 has been used to ignore selected charges from other rows
            info("Selected weight break profile " + str(weightbreak_profile_name))

            current_column_index = 1

            # Add offering type column at the beginning
            sheet.cell(row=1, column=current_column_index).value = "1"
            sheet.cell(row=2, column=current_column_index).value = "Offering Type"
            sheet.cell(row=3, column=current_column_index).value = "Offering Type"
            current_column_index += 1

            # Add rate type column as well
            sheet.cell(row=1, column=current_column_index).value = "2"
            sheet.cell(row=2, column=current_column_index).value = "Rate Type"
            sheet.cell(row=3, column=current_column_index).value = "Rate Type"
            current_column_index += 1
            
            # Populate column headers for all the static columns, set width for them and make them bold
            for column_index, template_static_column in enumerate(template_static_columns, start=current_column_index):
                sheet.cell(row=1, column=column_index).value = template_static_column.position
                sheet.cell(row=2, column=column_index).value = template_static_column.master_column_name
                sheet.cell(row=3, column=column_index).value = template_static_column.template_column_name

            current_column_index += len(template_static_columns)

            source_geo_hierarchy = ""
            dest_geo_hierarchy = ""
            
            # Get the values selected in dynamic parameters
            template_parameters = TemplateDynamicColumn.objects.filter(customer=customer, template=template, enabled="Y").order_by("position")
            for template_parameter in template_parameters:
                if template_parameter.parameter_name.startswith("Source"):
                    sheet.cell(row=1, column=current_column_index).value = template_parameter.position
                    sheet.cell(row=2, column=current_column_index).value = "SOURCE_GEO_HIERARCHY"
                    sheet.cell(row=3, column=current_column_index).value = "SOURCE_GEO_HIERARCHY"
                    current_column_index += 1
                elif template_parameter.parameter_name.startswith("Destination"):
                    sheet.cell(row=1, column=current_column_index).value = template_parameter.position
                    sheet.cell(row=2, column=current_column_index).value = "DEST_GEO_HIERARCHY"
                    sheet.cell(row=3, column=current_column_index).value = "DEST_GEO_HIERARCHY"
                    current_column_index += 1

                # Get the selected value for the dynamic parameter
                field_value = request.POST.get(template_parameter.html_element_name + "_0", "")  # index 0 has been used to ignore dynamic parameters from other rows
                if field_value != "":
                    # Get the columns that should be added for the given selected value
                    template_dynamic_column = TemplateDynamicColumnValue.objects.get(id = int(field_value), customer=customer)
                    if template_parameter.parameter_name.startswith("Source"):
                        source_geo_hierarchy = template_dynamic_column.parameter_value
                    elif template_parameter.parameter_name.startswith("Destination"):
                        dest_geo_hierarchy = template_dynamic_column.parameter_value
                    # Populate column headers for all the dynamic columns, set width for them and make them bold
                    index = 0
                    for column_index, template_column_name in enumerate(template_dynamic_column.get_template_columns(), start=current_column_index):
                        sheet.cell(row=1, column=column_index).value = float(str(template_parameter.position) + "." + str(index+1))
                        sheet.cell(row=2, column=column_index).value = template_column_name
                        sheet.cell(row=3, column=column_index).value = template_dynamic_column.get_master_columns()[index]
                        index += 1
                    
                    current_column_index += len(template_dynamic_column.get_template_columns())
                else:
                    field_name = ""
                    if template_parameter.parameter_name.startswith("Source"):
                        field_name = "source_region_0" # index 0 has been used to ignore dynamic parameters from other rows
                        field_value = request.POST.get(field_name, "")
                        if field_value != "":
                            sheet.cell(row=1, column=current_column_index).value = float(str(template_parameter.position) + ".1")
                            sheet.cell(row=2, column=current_column_index).value = "Source Region"
                            sheet.cell(row=3, column=current_column_index).value = "Source Region"
                            current_column_index += 1
                    elif template_parameter.parameter_name.startswith("Destination"):
                        field_name = "destination_region_0" # index 0 has been used to ignore dynamic parameters from other rows
                        field_value = request.POST.get(field_name, "")
                        if field_value != "":
                            sheet.cell(row=1, column=current_column_index).value = float(str(template_parameter.position) + ".1")
                            sheet.cell(row=2, column=current_column_index).value = "Destination Region"
                            sheet.cell(row=3, column=current_column_index).value = "Destination Region"
                            current_column_index += 1

            # Populate accessorial charges columns
            for column_index, charges in enumerate(selected_charges, start=current_column_index):
                sheet.cell(row=1, column=column_index).value = (utility.STARTING_POSITION_OF_ACCESSORIALS + column_index)
                sheet.cell(row=2, column=column_index).value = "accessorialCodeXid|" + charges
                sheet.cell(row=3, column=column_index).value = charges
                concatenated_charges += charges + "|"

            current_column_index += len(selected_charges)

            # Populate weight break related columns
            if weightbreak_profile_name != "":
                weight_breaks = get_weight_break(instance, weightbreak_profile_name)
                if len(weight_breaks) > 0:
                    column_index = current_column_index
                    for weightBreak in weight_breaks:
                        sheet.cell(row=1, column=column_index).value = (utility.STARTING_POSITION_OF_WEIGHTBREAKS + column_index)
                        sheet.cell(row=2, column=column_index).value = "weightBreak|" + weightBreak.get("WeightBreakGid")
                        sheet.cell(row=3, column=column_index).value = weightBreak.get("WeightBreak")
                        column_index += 1
                        sheet.cell(row=1, column=column_index).value = (utility.STARTING_POSITION_OF_WEIGHTBREAKS + column_index)
                        sheet.cell(row=2, column=column_index).value = "weightBreakMinCost|" + weightBreak.get("WeightBreakGid")
                        sheet.cell(row=3, column=column_index).value = "Min. Cost"
                        column_index += 1
                        sheet.cell(row=1, column=column_index).value = (utility.STARTING_POSITION_OF_WEIGHTBREAKS + column_index)
                        sheet.cell(row=2, column=column_index).value = "weightBreakMaxCost|" + weightBreak.get("WeightBreakGid")
                        sheet.cell(row=3, column=column_index).value = "Max. Cost"
                        column_index += 1
                        
                    current_column_index += len(weight_breaks) * 3
            
            column_names = [(sheet.cell(row=2, column=columnIndex).value) for columnIndex in range(1, sheet.max_column+1) if str(sheet.cell(row=2, column=columnIndex).value).strip != ""]
            
            # Set the default values in Offering Type column for the given rows
            mode = offering_type_description if offering_type_description != "" else offering_type_name
            column_name = "Offering Type"
            column_value = mode.upper()
            if column_value.strip() != "" and column_name in column_names:
                column_index = column_names.index(column_name) + 1
                for row_index in range(data_starts_from_row, int(record_count) + data_starts_from_row):
                    sheet.cell(row=row_index, column=column_index).value = column_value

            # Set the default values in Rate Type column for the given rows
            column_name = "Rate Type"
            column_value = rate_type_name.upper()
            if column_value.strip() != "" and column_name in column_names:
                column_index = column_names.index(column_name) + 1
                for row_index in range(data_starts_from_row, int(record_count) + data_starts_from_row):
                    sheet.cell(row=row_index, column=column_index).value = column_value

            # Set the default values in Service Provider Name, Service Provider ID and SCAC columns for the given rows based on Service Provider Name
            provider_name = request.POST.get("service_provider_name_0", "")
            if provider_name != "":
                column_name = "Service Provider Name"
                column_value = provider_name
                if column_value.strip() != "" and column_name in column_names:
                    column_index = column_names.index(column_name) + 1
                    for row_index in range(data_starts_from_row, int(record_count) + data_starts_from_row):
                        sheet.cell(row=row_index, column=column_index).value = column_value

            # Get service provider details from session including service provider id and scac
            providers = request.session["service_providers_details"]

            # Set provider id
            if provider_name != "" and provider_name in providers.keys():
                provider_id = providers[provider_name].get("id")
                column_name = "Service Provider ID"
                column_value = provider_id
                if column_value.strip() != "" and column_name in column_names:
                    column_index = column_names.index(column_name) + 1
                    for row_index in range(data_starts_from_row, int(record_count) + data_starts_from_row):
                        sheet.cell(row=row_index, column=column_index).value = column_value

            # Set SCAC
            if provider_name != "" and provider_name in providers.keys():
                scac = providers[provider_name].get("scac")
                column_name = "SCAC"
                column_value = scac
                if column_value.strip() != "" and column_name in column_names:
                    column_index = column_names.index(column_name) + 1
                    for row_index in range(data_starts_from_row, int(record_count) + data_starts_from_row):
                        sheet.cell(row=row_index, column=column_index).value = column_value
            
            # Set the default values for rest of the static columns
            for template_static_column in template_static_columns:
                field_value = request.POST.get(template_static_column.html_element_name + "_" + str(rowIndex), "")
                column_name = template_static_column.template_column_name
                column_value = field_value
                if column_value.strip() != "" and column_name in column_names:
                    column_index = column_names.index(column_name) + 1
                    for row_index in range(data_starts_from_row, int(record_count) + data_starts_from_row):
                        sheet.cell(row=row_index, column=column_index).value = column_value

            # Set the default values for the Source Geography/Region
            source_geography = request.POST.get("source_geography_0", "")
            source_region = request.POST.get("source_region_0", "")
            if source_geography != "":
                column_name = "SOURCE_GEO_HIERARCHY"
                column_value = source_geo_hierarchy
                if column_value.strip() != "" and column_name in column_names:
                    column_index = column_names.index(column_name) + 1
                    for row_index in range(data_starts_from_row, int(record_count) + data_starts_from_row):
                        sheet.cell(row=row_index, column=column_index).value = column_value 
            elif source_region != "":
                column_name = "Source Region"
                column_value = source_region
                if column_value.strip() != "" and column_name in column_names:
                    column_index = column_names.index(column_name) + 1
                    for row_index in range(data_starts_from_row, int(record_count) + data_starts_from_row):
                        sheet.cell(row=row_index, column=column_index).value = column_value

                column_name = "SOURCE_GEO_HIERARCHY"
                column_value = "Region"
                if column_value.strip() != "" and column_name in column_names:
                    column_index = column_names.index(column_name) + 1
                    for row_index in range(data_starts_from_row, int(record_count) + data_starts_from_row):
                        sheet.cell(row=row_index, column=column_index).value = column_value 

            # Set the default values for the Destination Geography/Region
            destination_geography = request.POST.get("destination_geography_0", "")
            destination_region = request.POST.get("destination_region_0", "")
            if destination_geography != "":
                column_name = "DEST_GEO_HIERARCHY"
                column_value = dest_geo_hierarchy
                if column_value.strip() != "" and column_name in column_names:
                    column_index = column_names.index(column_name) + 1
                    for row_index in range(data_starts_from_row, int(record_count) + data_starts_from_row):
                        sheet.cell(row=row_index, column=column_index).value = column_value 
            elif destination_region != "":
                column_name = "Destination Region"
                column_value = destination_region
                if column_value.strip() != "" and column_name in column_names:
                    column_index = column_names.index(column_name) + 1
                    for row_index in range(data_starts_from_row, int(record_count) + data_starts_from_row):
                        sheet.cell(row=row_index, column=column_index).value = column_value

                column_name = "DEST_GEO_HIERARCHY"
                column_value = "Region"
                if column_value.strip() != "" and column_name in column_names:
                    column_index = column_names.index(column_name) + 1
                    for row_index in range(data_starts_from_row, int(record_count) + data_starts_from_row):
                        sheet.cell(row=row_index, column=column_index).value = column_value          

            data_starts_from_row = int(record_count) + data_starts_from_row

        # Store the column index in a list so that they can be sorted
        column_seq_val =  [float(sheet.cell(row=1, column=columnIndex).value) for columnIndex in range(1, sheet.max_column+1) if str(sheet.cell(row=1, column=columnIndex).value).strip != "" ]
        source_column_index = [(column_seq_val.index(x) + 1) for x in sorted(column_seq_val)]
        info(str(source_column_index))
        
        # Create a new workbook which will have columns in correct order, but this Excel would not have the sequence row
        wb_new = Workbook()  
        sheet_new = wb_new.active
        # Copy the cells to this new workbook
        thin_border = Side(border_style="thin", color="000000")
        filled_cell_fill = PatternFill(start_color="E5E4E2", end_color="E5E4E2", fill_type = "solid")
        for r in range (2, sheet.max_row+1):
            for c in range (1, sheet.max_column+1):
                sheet_new.cell(row=r-1, column=c).value = sheet.cell(row=r, column=source_column_index[c-1]).value
                sheet_new.cell(row=r-1, column=c).border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
                if r > 2 and str(sheet_new.cell(row=r-1,column=c).value) != "" and sheet_new.cell(row=r-1, column=c).value is not None:
                    sheet_new.cell(row=r-1, column=c).fill = filled_cell_fill

        # Make the cells from row 2 as bold and font color as blue
        first_row_font = Font(bold=True, color="FF0000")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="152B61", end_color="152B61", fill_type = "solid")
        for c in range (1, sheet.max_column+1):
            cell = sheet_new.cell(row=1, column=c)  
            cell.font = first_row_font
            cell = sheet_new.cell(row=2, column=c)  
            cell.font = header_font
            cell.fill = header_fill
            sheet_new.column_dimensions[get_column_letter(c)].width = 30

        # Indentify column position of Contract Name, Start Date and End Date
        # contract_field_column_index = -1
        # start_date_column_index = -1
        # end_date_column_index = -1
        # for c in range (1, sheet.max_column+1):
        #     if sheet_new.cell(row=2, column=c).value == "Contract Name": contract_field_column_index = c
        #     if sheet_new.cell(row=2, column=c).value == "Start Date": start_date_column_index = c
        #     if sheet_new.cell(row=2, column=c).value == "End Date": end_date_column_index = c

        # Change the color of Start Date and End Date columns
        # filled_cell_fill = PatternFill(start_color="E5E4E2", end_color="E5E4E2", fill_type = "solid")
        # for r in range (3, sheet.max_row):
        #     sheet_new.cell(row=r, column=start_date_column_index).fill = filled_cell_fill
        #     sheet_new.cell(row=r, column=end_date_column_index).fill = filled_cell_fill

        # Delete some columns based on condition
        # existing_contract = request.POST.get("existing_contract_0", "")
        # contractName = request.POST.get("contract_name_0", "")
        # if existing_contract == "Y" and contractName != "":
        #     if start_date_column_index != -1: sheet_new.delete_cols(start_date_column_index)
        #     if end_date_column_index != -1: sheet_new.delete_cols(end_date_column_index)
        # else:
        #     if contract_field_column_index != -1: sheet_new.delete_cols(contract_field_column_index)

        # Hide some columns
        all_column_names = [(sheet_new.cell(row=1, column=columnIndex).value) for columnIndex in range(1, sheet_new.max_column+1) if str(sheet_new.cell(row=2, column=columnIndex).value).strip != ""]
        columns_to_be_hidden = ["Offering Type", "Rate Type", "Service Provider ID", "SCAC", "SOURCE_GEO_HIERARCHY", "DEST_GEO_HIERARCHY"]
        for column_to_be_hidden in columns_to_be_hidden:
            if column_to_be_hidden in all_column_names:
                column_index = all_column_names.index(column_to_be_hidden) + 1
                sheet_new.column_dimensions[get_column_letter(column_index)].hidden= True
        
        # Hide first row
        sheet_new.row_dimensions[1].hidden = True
        
        # Freeze header rows
        sheet_new.freeze_panes = sheet_new['A3'] # Row 3 has been used because the 1st row will be hidden
        
        # Enable protection for the whole sheet
        #sheet.protection.sheet = True

        # Disable protection for rows except header rows
        # for row_cells in sheet.iter_rows(min_row=3):
        #     for cell in row_cells:
        #         cell.protection = Protection(locked=False)        
        
        # content-type of response

        # Save template to DB, if save template checkbox is enabled
        save_template = request.POST.get("save_template_cb", "NO") 
        if save_template == "YES":
            batch_template_name = request.POST.get("template_name", "") 
            if batch_template_name == "": batch_template_name = "Template"
            batch_template_name += "-" + template.offering_type_name + "-" + template.rate_type_name.replace(" ", "_") + "-" + datetime.today().strftime('%Y%m%d%H%M%S')
            try:
                concatenated_charges = concatenated_charges.strip("|") # Selected charges will be concatenated by pipe before storing 
                info("concatenated_charges = " + str(concatenated_charges))
                temp_request_dict = dict(request.POST.items())
                temp_request_dict["charges_0"] = concatenated_charges
                temp_request_dict.pop('csrfmiddlewaretoken')
                batch = Batch(customer = customer, template = template, batch_name = batch_template_name, field_values = json.dumps(temp_request_dict), created_by = request.user.username)
                batch.save()
            except Exception as ex:
                error("generate_template_generic - " + str(ex))

        response = HttpResponse(content_type='application/ms-excel')

        #decide file name
        response['Content-Disposition'] = 'attachment; filename="'+template.offering_type_name + "-" + template.rate_type_name.replace(" ", "_") + "-" + datetime.today().strftime('%Y%m%d%H%M%S') + '.xlsx"'

        wb_new.save(response)
        #wb.save(response)
        return response

    else:
        customer_key = ""
        # Get customer key from the session and fetch customer object
        if "customer_key" in request.session:
            customer_key = request.session["customer_key"]
        customer = Customer.objects.get(customer_key=customer_key)

        customer_key = ""
        template_id = 0
        offering_type_name = ""
        offering_type_description = ""
        rate_type_name = ""

        # Get values from session
        if "customer_key" in request.session:
            customer_key = request.session["customer_key"]
        if "template_id" in request.session:
            template_id = int(request.session["template_id"])
        if "offering_type_name" in request.session:
            offering_type_name = request.session["offering_type_name"]
        if "offering_type_description" in request.session:
            offering_type_description = request.session["offering_type_description"]
        if "rate_type_name" in request.session:
            rate_type_name = request.session["rate_type_name"]

        info("customer_key " + customer_key)
        info("template_id " + str(template_id))
        info("offering_type_name " + offering_type_name)
        info("offering_type_description " + offering_type_description)
        info("rate_type_name " + rate_type_name)

        template_batch_id = 0
        try:
            if "template_batch_id" in request.session:
                if request.session["template_batch_id"] != "":
                    template_batch_id = int(request.session["template_batch_id"]) # Store batch id in session
        except:
            error("generate_template_generic - Unable to find template_batch_id in session")

        info("template_batch_id = " + str(template_batch_id))  

        # Get template object
        template = Template.objects.get(id=template_id, customer=customer)        

        # Get all the selectable static columns 
        template_static_columns = TemplateStaticColumn.objects.filter(customer=customer, template=template, enterable="Y", enabled="Y").order_by("position")
        if template_static_columns.count() == 0:
            return render(request, "rate/common/error_page.html", {"error_message" : "No static columns found for the template! Please contact the administrator."})

        # Get the parameters from dynamic columns
        template_parameters = TemplateDynamicColumn.objects.filter(customer=customer, template=template, enabled="Y").order_by("position")

        # Get the parameter values from dynamic columns
        template_parameter_values = TemplateDynamicColumnValue.objects.filter(customer=customer, template=template, enabled="Y").order_by("sequence")

        # Get all curency codes
        currencies = Currency.objects.all()
        info("Count of currencies is " + str(len(currencies)))

        # Get instance and domain from session
        instance_id = 0
        domain = ""
        if "instance_id" in request.session:
            instance_id = int(request.session["instance_id"])
        if "domain" in request.session:
            domain = request.session["domain"]

        instance = Instance.objects.get(id=instance_id, customer=customer)

        # Get list of service providers
        service_providers, service_providers_details = get_service_providers(instance, domain)

        # Store service provider details in session which includes service provider id and scac
        request.session["service_providers_details"] = service_providers_details

        # Get list of regions
        regions = get_regions(instance)

        volume_uoms = []
        weight_uoms = []
        equipments = []
        acs_charges = []
        weight_break_profiles = []

        if rate_type_name.upper() != "DISCOUNT":
            # Get list of volume UOMs
            volume_uoms = UOM.objects.filter(uom_type = "VOLUME")

            # Get list of weight UOMs
            weight_uoms = UOM.objects.filter(uom_type = "WEIGHT")

            # Get list of equipments
            equipments = get_equipments(instance)

            # Get list of accessorial costs
            acs_charges = get_acessorial_codes(instance)

            # Get list of unit break profiles
            weight_break_profiles = get_unit_break_profiles(instance, "WEIGHT")

        # Template batch fields, if template was selected
        template_batch_fields = ""
        if template_batch_id != 0:
            batch = Batch.objects.get(id = template_batch_id)
            template_batch_fields = batch.field_values

        info("template_batch_fields " + template_batch_fields)

        # Show download template page
        return render(request, "rate/customer/user/generate_template_generic.html", {
            "template_id" : template_id, 
            "template_static_columns" : template_static_columns,
            "template_parameters" : template_parameters, 
            "template_parameter_values" : template_parameter_values,
            "currencies" : currencies,
            "service_providers" : service_providers,
            "equipments" : equipments,
            "acs_charges" : acs_charges,
            "regions" : regions,
            "volume_uoms" : volume_uoms,
            "weight_uoms" : weight_uoms,
            "weight_break_profiles" : weight_break_profiles,
            "template_batch_fields" : template_batch_fields,
            "show_add_remove_buttons" : utility.SHOW_ADD_REMOVE_BUTTONS,
            "starting_position_of_accessorials" : utility.STARTING_POSITION_OF_ACCESSORIALS,
            "starting_position_of_weightbreaks" : utility.STARTING_POSITION_OF_WEIGHTBREAKS
        } 
    )


@login_required
def upload_template(request, key):
    info("========In upload_template========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    # If the user is not a staff user, and if the user is customer ADMIN but is not suppsed to have access of rate upload screens, then show error
    if request.user.is_staff == False:
        if request.user.profile.user_type == "ADMIN" and request.session["allow_admin_rate_upload"] == "N":
            return render(request, "rate/common/error_page.html", {"error_message" : "ADMIN user is not authorized to view this page."})

    if request.method == 'POST' and len(list(request.FILES.items())) > 0:
        customer_key = ""
        template_id = 0
        instance_id = 0
        offering_type_name = ""
        offering_type_description = ""
        rate_type_name = ""
        result = ""
        
        if "customer_key" in request.session:
            customer_key = request.session["customer_key"]
        if "template_id" in request.session:
            template_id = int(request.session["template_id"])
        if "instance_id" in request.session:
            instance_id = int(request.session["instance_id"])
        if "offering_type_name" in request.session:
            offering_type_name = request.session["offering_type_name"]
        if "offering_type_description" in request.session:
            offering_type_description = request.session["offering_type_description"]
        if "rate_type_name" in request.session:
            rate_type_name = request.session["rate_type_name"]

        # Save current timestaml in session, this will be appended to the uplaoded Excel template file name and all other output files
        timestamp = datetime.today().strftime('%Y%m%d%H%M%S')
        request.session["timestamp"] = timestamp

        info("customer_key " + customer_key)
        info("template_id " + str(template_id))
        info("offering_type_name " + offering_type_name)
        info("offering_type_description " + offering_type_description)
        info("rate_type_name " + rate_type_name)

        # Get profile 
        profile = request.user.profile

        # Get customer 
        customer = Customer.objects.get(customer_key=customer_key)

        # Get instance
        instance = Instance.objects.get(id=instance_id, customer=customer)

        # Create directory where the excel template will be uploaded and the CSVs will be created
        # This directory will be ./media/<Customer ID>/<Instance Name> (Non-alphanumeric characters will be removed from instance name)
        csv_dir = os.path.join(settings.MEDIA_ROOT, str(customer.customer_key), re.sub(r'[^a-zA-Z0-9]', '', instance.instance_name))
        if not os.path.exists(csv_dir):
            os.makedirs(csv_dir)       

        # Get the uploaded file from request and create an instance of file system storage
        rate_file = request.FILES['ratefile']
        fs = FileSystemStorage(location=csv_dir)

        # Get the file name and extension of the uploaded file
        split_tup = os.path.splitext(rate_file.name)

        # Append timestamp and create a new file name
        rate_file_name = split_tup[0] + "_" + timestamp + split_tup[1]

        # Save the uploaded Excel
        file_name = fs.save(rate_file_name, rate_file)

        # Get the full path of the saved Excel
        excel_file_path = os.path.join(fs.location, file_name)

        # Initialize an empty dataframe to store the uploaded Excel
        input_df = pd.DataFrame()

        # Use different method to read the excel based on the its extension
        if file_name.lower().endswith(".xlsx"):
            # Open the xlsx template and store it as pandas data frame
            input_df = pd.read_excel(excel_file_path, engine='openpyxl', dtype="string")
        elif file_name.lower().endswith(".xls"):
            # Open the xls template and store it as pandas data frame
            input_df = pd.read_excel(excel_file_path, dtype="string")
        else:
            error("Invalid file format! Only .xlsx or .xls files are allowed.")
            return render(request, "rate/common/error_page.html", {"error_message" : "Invalid file format! Only .xlsx or .xls files are allowed."})

        # Remove the 2nd row of the data frame (which is dummy template column header) and reset index of the dataframe
        input_df = input_df.drop(0).reset_index()

         # Convert all fields to upper case if the datatype is string
        for column in input_df.columns:
            try:
                if input_df[column].dtypes == "string":
                    input_df[column] = input_df[column].astype(str).str.upper()
            except Exception as e:
                error("upload_template - Error while converting input data to upper case " + str(e))

        # Drop rows if Charge Amount is missing
        input_df.drop(input_df[(input_df["Charge Amount"] == "<NA>")].index, inplace=True)

        # Check if Charge Amount contains non-numeric values
        for charge_amount in input_df["Charge Amount"]:
            try:
                float(charge_amount)                
            except:
                error("Charge Amount column contains non-numeric data")
                return render(request, "rate/common/error_page.html", {"error_message" : "Charge Amount column contains non-numeric data - e.g., " + str(charge_amount)})
        
        if "Start Date" in input_df.columns.values.tolist():
            # Check if Start Date contains date in correct format, if no show error, if yes update input_df with YYYYMMDD format
            start_date_formatted = []
            for start_date in input_df["Start Date"]:
                try:
                    if start_date.strip() != "":
                        start_date_str = start_date.strip().split(" ")[0]
                        start_date_dt = parser.parse(start_date_str, yearfirst = True, dayfirst = False)
                        print("start_date_dt")
                        print(start_date_dt)
                        start_date_formatted.append(start_date_dt.strftime('%Y%m%d') + "000000")
                        print(start_date_dt.strftime('%Y%m%d'))
                        print(start_date_dt.strftime('%Y%m%d') + "000000")
                    else:
                        start_date_formatted.append("")
                except:
                    error("Start Date column contains invalid dates")
                    return render(request, "rate/common/error_page.html", {"error_message" : f"Start Date column contains invalid dates - {start_date} (date format should be YYYY/MM/DD or YYYY-MM-DD or YYYY.MM.DD or YYYYMMDD or YYYY-MMM-DD).\
                    Example: 2010/08/15, 2010-08-15, 2010.08.15, 20100815, 2010-Aug-15 etc."})
            
            input_df["Start Date"] = start_date_formatted

        if "End Date" in input_df.columns.values.tolist():
            # Check if End Date contains date in correct format, if no show error, if yes update input_df with YYYYMMDD format
            end_date_formatted = []
            for end_date in input_df["End Date"]:
                try:
                    if end_date.strip() != "":
                        end_date_str = end_date.strip().split(" ")[0]
                        end_date_dt = parser.parse(end_date_str, yearfirst = True, dayfirst = False)
                        end_date_formatted.append(end_date_dt.strftime('%Y%m%d') + "000000")
                    else:
                        end_date_formatted.append("")
                except:
                    error("End Date column contains invalid dates")
                    return render(request, "rate/common/error_page.html", {"error_message" : f"End Date column contains invalid dates - {end_date} (date format should be YYYY/MM/DD or YYYY-MM-DD or YYYY.MM.DD or YYYYMMDD or YYYY-MMM-DD).\
                    Example: 2010/08/15, 2010-08-15, 2010.08.15, 20100815, 2010-Aug-15 etc."})
            
            input_df["End Date"] = end_date_formatted
        
        info("Printing input_df...")
        info(str(input_df))

        # If dataframe contains no rows then show error
        if len(input_df) == 0:
            error("Uploaded Excel contains no valid rows")
            return render(request, "rate/common/error_page.html", {"error_message" : "Uploaded Excel contains no valid rows."})

        # Get column names of the data frame
        input_column_headers = input_df.columns.values.tolist()
        info("Column names of the uploaded file:")
        info(str(input_column_headers))

        # Get the list of columns that should be checked for duplicates
        column_headers_for_duplicates = input_column_headers.copy()
        column_headers_for_duplicates.remove("index") # index column should not be considered while checking duplicates
        column_headers_for_duplicates.remove("Charge Amount") # Charge Amount column should not be considered while checking duplicates
        if "Service Time Value" in column_headers_for_duplicates:
            column_headers_for_duplicates.remove("Service Time Value") # Service Time Value column should not be considered while checking duplicates
        if "Min Charge" in column_headers_for_duplicates:
            column_headers_for_duplicates.remove("Min Charge") # Min Charge column should not be considered while checking duplicates
        info("Column Headers for Duplicates ")
        info(str(column_headers_for_duplicates))
        duplicate = input_df[input_df.duplicated(column_headers_for_duplicates, keep=False)] # Get the duplicate records
        if len(duplicate) > 0:
            duplicate_row_nums = duplicate["index"].tolist()
            duplicate_row_nums = [(int(i)+2) for i in duplicate_row_nums] # Actual row number in the Excel will be 2 more
            info("Duplicate row numbers " + str(duplicate_row_nums))
            return render(request, "rate/common/error_page.html", {"error_message" : "Duplicate records found in the uploaded Excel. Please check following rows.", "duplicate_row_nums" : duplicate_row_nums})

        # Declare empty list to contain the list of CSVs that will be generated 
        csv_file_list = []
        # Get the list of CSV files that are supposed to be generated for the customer
        csv_files = CsvFile.objects.filter(customer=customer, enabled="Y").order_by("sequence")
        info("CSV file count " + str(len(csv_files)))
        if csv_files.count() == 0:
            return render(request, "rate/common/error_page.html", {"error_message" : "No CSV file has been defined! Please contact the administrator."})

        # Set common variables which will written to the CSV, some of these will be used in the expression field of the CsvStructure data model
        DOMAIN = str(request.session["domain"]).upper()
        YEAR = str(date.today().year)
        YMD = str(datetime.now().strftime("%y%m%d")).upper()
        OFFERING_TYPE = offering_type_name.upper()
        RATE_TYPE = rate_type_name.upper()
        REGION = ""
        if profile.region:
            REGION = profile.region

        EXISTING_CONTRACT_EXISTS = any([(input_column_header.strip() == "Contract Name") for input_column_header in input_column_headers])
        info("EXISTING_CONTRACT_EXISTS " + str(EXISTING_CONTRACT_EXISTS))

        # Store all attributes in session, these are template specific and can be used during CSV generation
        ATTRIBUTE1 = str(request.session["attribute1"]) 
        ATTRIBUTE2 = str(request.session["attribute2"]) 
        ATTRIBUTE3 = str(request.session["attribute3"]) 
        ATTRIBUTE4 = str(request.session["attribute4"]) 
        ATTRIBUTE5 = str(request.session["attribute5"]) 
        ATTRIBUTE6 = str(request.session["attribute6"]) 
        ATTRIBUTE7 = str(request.session["attribute7"]) 
        ATTRIBUTE8 = str(request.session["attribute8"]) 
        ATTRIBUTE9 = str(request.session["attribute9"]) 
        ATTRIBUTE10 = str(request.session["attribute10"])

        info("Region from profile " + REGION)
        info("Attributes from session " + ",".join([ATTRIBUTE1, ATTRIBUTE2, ATTRIBUTE3, ATTRIBUTE4, ATTRIBUTE5, ATTRIBUTE6, ATTRIBUTE7, ATTRIBUTE8, ATTRIBUTE9, ATTRIBUTE10]))
        
        # Get the timestamp from session, this will be appended to the CSV file name
        timestamp = request.session["timestamp"]
        # Loop over each CSV file
        for csv_file in csv_files: 
            info("--------Processing CSV " + csv_file.name + "--------")
            # Get all the fields of the given CSV file
            csv_fields = CsvStructure.objects.filter(csv_file=csv_file, enabled="Y").order_by("sequence")
            if csv_fields.count() == 0:
                #return render(request, "rate/common/error_page.html", {"error_message" : "No field for the CSV file has been defined! Please contact the administrator."})
                info("No field for the CSV file has been defined, skipping this file")
                continue
         
            # Get array of CSV column names 
            output_column_headers = [(csv_field.column_name) for csv_field in csv_fields]
            info("Column names of the output file:")
            info(str(output_column_headers))

            # Prepare the output data frame including the column names and blank rows, row count should be same as the input df
            output_df = pd.DataFrame(columns = output_column_headers, index=range(len(input_df)))

            for csv_field in csv_fields:
                if csv_field.fixed_value.strip() != "":
                    # If fixed value has been provided, then hard code
                    info("Populating output column " + csv_field.column_name + " as " + csv_field.fixed_value.strip())
                    output_df[csv_field.column_name] = csv_field.fixed_value.strip()
                elif csv_field.source_column_name.strip() != "":
                    # If source has been provided, then take it from input df
                    info("Populating output column " + csv_field.column_name + " from " + csv_field.source_column_name.strip() + " column of input file")
                    if csv_field.source_column_name.strip() in input_df.columns:                        
                        output_df[csv_field.column_name] = input_df[csv_field.source_column_name.strip()]
                    else:
                        info("Column name " + csv_field.column_name + " not found in input file")
                elif csv_field.expression.strip() != "":
                    # If expression has been provided, then execute that
                    try:
                        info("Populating output column " + csv_field.column_name + " using expression " + csv_field.expression.strip())
                        output_df[csv_field.column_name] =  eval(csv_field.expression.strip())
                        # If the column is a GID or XID column, then take only first 50 characters as some GID/XID columns have 50 characters length
                        if csv_field.column_name.endswith("_GID") or csv_field.column_name.endswith("_XID"):
                            output_df[csv_field.column_name] = output_df[csv_field.column_name].str[0:50]
                    except Exception as e:
                        error("Error while executing expression for " + csv_field.column_name + "-" + str(e))
                else:
                    # If source of the columns has NOT been mentioned in the applicaiton, then derive it here
                    if csv_file.name == "RATE_GEO" and csv_field.column_name == "RATE_OFFERING_GID":
                        output_df[csv_field.column_name] = input_df.apply(derive_rate_offering_gid, args=(DOMAIN, EXISTING_CONTRACT_EXISTS, YMD, REGION), axis = 1)

            # Populate dataframe for accessorials cost
            if csv_file.name == "ACCESSORIAL_COST":
                # Get all the accessorial related column names from input Excel
                ass_columns = [input_column_header for input_column_header in input_column_headers if input_column_header.startswith("accessorialCodeXid|")]
                if len(ass_columns) == 0:
                    # If no accessorial columns are there in Excel, remove all rows from the dataframe so that the CSV file doesn't get created
                    info("No accessorial cost columns found in the Excel.")
                    output_df.drop(output_df.index, inplace=True)
                else:
                    # Append all accessorial costs into a pipe delimited value and store in in Charge Amount column of the output_df
                    output_df["CHARGE_AMOUNT"] = ""
                    for ass_column in ass_columns:
                        output_df["CHARGE_AMOUNT"] = output_df["CHARGE_AMOUNT"] + "|" + input_df[ass_column]

                    # Drop all the duplicates from output dataframe and create a new dataframe with it
                    temp_output_df = output_df.drop_duplicates()
                    # Delete all rows from output_df
                    output_df = output_df[0:0]
                    # Iterate over all rows
                    for row_index, row in temp_output_df.iterrows():
                        ass_costs = str(row["CHARGE_AMOUNT"]).strip("|")
                        # Iterate over accessorial columns
                        for ass_column_index, ass_column in enumerate(ass_columns):
                            ass_code = ass_column.replace("accessorialCodeXid|", "").strip()
                            ass_cost = (ass_costs.split("|")[ass_column_index]).strip()
                            if ass_cost != "" and ass_cost != "<NA>":
                                try:
                                    ass_cost_float = float(ass_cost)
                                except:
                                    # Show error if any non-numeric value was provided as accessorial cost
                                    error(f"{ass_column.replace('accessorialCodeXid|', '')} column contains non-numeric data")
                                    return render(request, "rate/common/error_page.html", {"error_message" : f"{ass_column.replace('accessorialCodeXid|', '')} column contains non-numeric data - e.g., " + ass_cost})

                                row["ACCESSORIAL_COST_GID"] = row["ACCESSORIAL_COST_GID"] + "-" + ass_code
                                row["ACCESSORIAL_COST_XID"] = row["ACCESSORIAL_COST_XID"] + "-" + ass_code
                                row["CHARGE_AMOUNT"] = ass_cost
                                # Append row to output_df
                                # output_df = output_df.append(row, ignore_index=True)  # working correctly
                                # output_df = pd.concat([output_df, row], ignore_index=True, sort=False)
                                # at_row = len(output_df)
                                # output_df.iloc[at_row] = row
                                output_df = pd.concat([output_df, row.to_frame().T])

            # Populate dataframe for accessorials
            if csv_file.name == "RATE_GEO_ACCESSORIAL":
                # Get all the accessorial related column names from input Excel
                ass_columns = [input_column_header for input_column_header in input_column_headers if input_column_header.startswith("accessorialCodeXid|")]
                if len(ass_columns) == 0:
                    # If no accessorial columns are there in Excel, remove all rows from the dataframe so that the CSV file doesn't get created
                    info("No accessorial cost columns found in the Excel.")
                    output_df.drop(output_df.index, inplace=True)
                else:
                    # Append all accessorial costs into a pipe delimited value and store in in Charge Amount column of the output_df
                    output_df["CHARGE_AMOUNT"] = ""
                    for ass_column in ass_columns:
                        output_df["CHARGE_AMOUNT"] = output_df["CHARGE_AMOUNT"] + "|" + input_df[ass_column]

                    # Drop all the duplicates from output dataframe and create a new dataframe with it
                    temp_output_df = output_df.drop_duplicates()
                    # Delete all rows from output_df
                    output_df = output_df[0:0]
                    # Iterate over all rows
                    for row_index, row in temp_output_df.iterrows():
                        ass_costs = str(row["CHARGE_AMOUNT"]).strip("|")
                        # Iterate over accessorial columns
                        for ass_column_index, ass_column in enumerate(ass_columns):
                            ass_code = ass_column.replace("accessorialCodeXid|", "").strip()
                            ass_cost = (ass_costs.split("|")[ass_column_index]).strip()
                            if ass_cost != "" and ass_cost != "<NA>":
                                try:
                                    ass_cost_float = float(ass_cost)
                                except:
                                    # Show error if any non-numeric value was provided as accessorial cost
                                    error(f"{ass_column.replace('accessorialCodeXid|', '')} column contains non-numeric data")
                                    return render(request, "rate/common/error_page.html", {"error_message" : f"{ass_column.replace('accessorialCodeXid|', '')} column contains non-numeric data - e.g., " + ass_cost})

                                row["ACCESSORIAL_COST_GID"] = row["ACCESSORIAL_COST_GID"] + "-" + ass_code
                                row["ACCESSORIAL_CODE_GID"] = DOMAIN + "." + ass_code
                                # Append row to output_df
                                # output_df = output_df.append(row, ignore_index=True) # working correctly
                                # output_df = pd.concat([output_df, row], ignore_index=True, sort=False)
                                # at_row = len(output_df)
                                # output_df.iloc[at_row] = row
                                output_df = pd.concat([output_df, row.to_frame().T])

                    # Drop Charge Amount column which is not needed
                    output_df.drop("CHARGE_AMOUNT", axis=1, inplace=True)

            
            # If domain name is blank that means this is for existing contract
            output_df.drop(output_df[(output_df["DOMAIN_NAME"] == "")].index, inplace=True)

            # Remove duplicate records as per the setup
            if csv_file.auto_remove_duplicates == "Y":
                output_df = output_df.drop_duplicates()

            # Populate auto number field wherever applicable
            if csv_file.name in ["RATE_GEO_COST_GROUP", "RATE_GEO_COST"]:
                initial_sequence = 0
                if csv_file.name == "RATE_GEO_COST_GROUP":
                    initial_sequence = max_cost_group_seq(instance)
                    if initial_sequence == 0:
                        initial_sequence = csv_file.auto_number_initial_sequence if csv_file.auto_number_initial_sequence is not None else 99999
                    output_df["RATE_GEO_COST_GROUP_SEQ"] = range(initial_sequence, initial_sequence + len(output_df))
                if csv_file.name == "RATE_GEO_COST":
                    initial_sequence = max_cost_seq(instance)
                    if initial_sequence == 0:
                        initial_sequence = csv_file.auto_number_initial_sequence if csv_file.auto_number_initial_sequence is not None else 99999
                    output_df["RATE_GEO_COST_SEQ"] = range(initial_sequence, initial_sequence + len(output_df))

                # Update the auto_number_initial_sequence field with the last used value
                csv_file.auto_number_initial_sequence = initial_sequence + len(output_df) + 1
                csv_file.save(update_fields=['auto_number_initial_sequence'])

            # If the file name already include .csv, then remove it and then append .csv along with timestamp
            csv_file_name = re.sub('.csv$', '', csv_file.csv_file_name, flags=re.IGNORECASE) + "_" + timestamp + ".csv"
            csv_full_name = os.path.join(csv_dir, csv_file_name)

            info("Printing output_df...")
            info(str(output_df))

            try:
                # Write the dataframe to CSV without the index
                if len(output_df) > 0:
                    output_df.to_csv(csv_full_name, index=False, mode='w+')   

                    # Add lines to the file
                    first_line = csv_file.file_identifier.strip()
                    third_line = ""
                    if csv_file.has_date_field == "Y": 
                        third_line = "EXEC SQL ALTER SESSION SET NLS_DATE_FORMAT = '" + csv_file.date_format + "'"
                    if first_line != "":
                        dummy_file = csv_full_name + '.bak'
                        # open original file in read mode and dummy file in write mode
                        with open(csv_full_name, 'r') as read_obj, open(dummy_file, 'w') as write_obj:
                            # Write first line to the dummy file
                            write_obj.write(first_line + "\n")
                            # Write second line (column header) to the dummy file
                            write_obj.write(read_obj.readline())
                            if third_line != "":                 
                                # Write second line to the dummy file
                                write_obj.write(third_line + "\n")
                            # Read rest of the lines from original file one by one and append them to the dummy file
                            for line in read_obj:
                                write_obj.write(line)
                        # remove original file
                        os.remove(csv_full_name)
                        # Rename dummy file as the original file
                        os.rename(dummy_file, csv_full_name)

                    # Add the file to the list only if the dataframe has data
                    csv_file_list.append({"sequence" : csv_file.sequence, "file_path" : os.path.join(str(customer.customer_key), re.sub(r'[^a-zA-Z0-9]', '', instance.instance_name), csv_file_name), "file_name" : csv_file_name, "file_id" : csv_file.id})
                    info("Added CSV to the list " + csv_file.name)
                else:
                    info("Dataframe doesn't have any rows, no file will be written")
            except Exception as ex:
                 return render(request, "rate/common/error_page.html", {"error_message" : "Error while writing CSV files " + str(ex)})  
            
            info("CSV generated at " + csv_full_name)

        info(str(csv_file_list))

        # If csv files were created then zip them
        zip_file_path = ""
        if len(csv_file_list) > 0:
            zip_file_name = "ALL_FILES_" + timestamp + ".zip"
            zip_file_path = os.path.join(str(customer.customer_key), re.sub(r'[^a-zA-Z0-9]', '', instance.instance_name), zip_file_name)
            zip_file_full_name = os.path.join(csv_dir, zip_file_name)
            # Create the zip file in append mode
            zip_file = ZipFile(zip_file_full_name, "w")
            try:
                # Loop through all CSV files
                for csv_file in csv_file_list:
                    csv_file_name = csv_file.get("file_name", "")
                    if csv_file_name != "":
                        # Get the actual path of the CSV file
                        csv_file_name = os.path.join(csv_dir, csv_file_name)
                        # Add the CSV file to the zip
                        zip_file.write(csv_file_name, os.path.basename(csv_file_name), compress_type = ZIP_DEFLATED)
            except Exception as e:
                error("upload_template - error while creating zip file " + str(e))
            finally:
                # Close the zip file
                zip_file.close()

        result = "SUCCESS"

        info("Filename of the uploaded Excel " + file_name)

        return render(request, 'rate/customer/user/template_uploaded_page.html', {
            "input_file" : os.path.join(str(customer.customer_key), re.sub(r'[^a-zA-Z0-9]', '', instance.instance_name), file_name),
            "generated_csv_files" : csv_file_list,
            "zip_file" : zip_file_path,
            "result" : result,
            "instance" : instance
        })
    else:
        customer_key = ""
        # Get values from session
        customer_key = ""
        template_id = 0
        template_name = ""
        instance_id = 0

        # Get values from session
        if "customer_key" in request.session:
            customer_key = request.session["customer_key"]
        if "template_id" in request.session:
            template_id = int(request.session["template_id"])
        if 'template_name' in request.session:
            template_name = request.session['template_name']
        if "instance_id" in request.session:
            instance_id = int(request.session["instance_id"])

        # Get customer 
        customer = Customer.objects.get(customer_key=customer_key)

        if "instance_id" in request.session:
            instance_id = int(request.session["instance_id"])

        # Get instance
        instance = Instance.objects.get(id=instance_id, customer=customer)

        return render(request, 'rate/customer/user/rate_upload_page.html', {"instance" : instance})

@login_required
def csv_uploaded(request, key):
    info("========In csv_uploaded========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    # If the user is not a staff user, and if the user is customer ADMIN but is not suppsed to have access of rate upload screens, then show error
    if request.user.is_staff == False:
        if request.user.profile.user_type == "ADMIN" and request.session["allow_admin_rate_upload"] == "N":
            return render(request, "rate/common/error_page.html", {"error_message" : "ADMIN user is not authorized to view this page."})
        
    customer_key = ""
    instance_id = -1
    if "customer_key" in request.session:
        customer_key = request.session["customer_key"]
    if "instance_id" in request.session:
        instance_id = int(request.session["instance_id"])

    # Get customer 
    customer = Customer.objects.get(customer_key=customer_key)

    # Get instance
    instance = Instance.objects.get(id=instance_id, customer=customer)

    if request.method == "POST" and "GetStatus" in request.POST:
        upload_outcome = request.POST.get("upload_outcome")
        upload_message = request.POST.get("upload_message")
        transmission_no = request.POST.get("transmission_no")
        # Check upload result
        processing_status, processing_messages = get_transmission_status(instance, transmission_no)

        if processing_status == "PROCESSING":
            completed = "N"
        else:
            completed = "Y"

        return render(request, "rate/customer/user/csv_upload_result_page.html", {
            "outcome" : upload_outcome,
            "message" : upload_message,
            "transaction_id" : transmission_no,
            "processing_status" : processing_status,
            "processing_messages" : processing_messages,
            "completed" : completed
        })
    else:
        # Get the list of CSV file names that were generated
        csv_file_names = request.POST.getlist("csv_file_name[]")

        # Get the list of respective csv file ids
        csv_file_ids = request.POST.getlist("csv_file_id[]")

        info(str(csv_file_names))

        # Derive the CSv file path
        csv_dir = os.path.join(settings.MEDIA_ROOT, str(customer.customer_key), re.sub(r'[^a-zA-Z0-9]', '', instance.instance_name))

        # Get the list of csv_commands (iu or i) for each csv file
        csv_commands = []
        for csv_file_id in csv_file_ids:
            csv_file = CsvFile.objects.get(id=csv_file_id)
            csv_commands.append(csv_file.csv_command.lower())

        # Generate xml payload for all CSV files
        status, message, payload = generate_xml_payload(csv_dir, csv_file_names, csv_commands)

        # Delete the xml payload file
        if utility.AUTO_DELETE_OUTPUT_XML_FILE:
            try:
                for file in os.listdir(csv_dir):
                    if file.lower().endswith(".xml") and request.session["timestamp"] in file:
                        os.remove(os.path.join(csv_dir, file))
            except Exception as ex:
                error("Error while deleting output xml - " + str(ex))

        # Check if error was returned, if yes then show generic error page
        if status == "ERROR":
            return render(request, "rate/common/error_page.html", {
                "error_message" : message
            })

        # Upload payload to OTM
        status, message, transmission_no = push_rate_to_otm(instance, payload)

        # Check if error was returned by OTM, if yes then show that error 
        if status == "ERROR":
            return render(request, "rate/customer/user/csv_upload_result_page.html", {
                "outcome" : status,
                "message" : message,
                "completed" : "N"
            })

        # If everything was uploaded successfully, then show transmission no
        return render(request, "rate/customer/user/csv_upload_result_page.html", {
            "outcome" : status,
            "message" : message,
            "transaction_id" : transmission_no,
            "completed" : "N"
        })
    
    
@login_required
def manager_dashboard(request, key):
    info("========In manager_dashboard========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    # If the user is not a staff user or is not a customer user, then show error
    if request.user.is_staff == False:
        if request.user.profile.user_type != "ADMIN":
            return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    customer_key = ""
    if "customer_key" in request.session:
        customer_key = request.session["customer_key"]
    
    # profiles, instances, accesses, templates, batches for the current customer
    customer = Customer.objects.get(customer_key=customer_key)
    profiles = Profile.objects.filter(customer=customer).order_by("creation_date")
    instances = Instance.objects.filter(customer=customer).order_by("creation_date")
    accesses = UserAccess.objects.filter(customer=customer).order_by("creation_date")
    templates = Template.objects.filter(customer=customer).order_by("offering_type_name", "group_name", "rate_type_name")
    batches = Batch.objects.filter(customer=customer).order_by("-creation_date")

    return render(request, "rate/customer/superuser/manager_dashboard_page.html", {"current_user_name" : request.user.username, "profiles" : profiles, "instances" : instances, "accesses" : accesses, "templates" : templates, "batches" : batches})

def logout_view(request, key=""):
    info("========In logout_view========")
    customer_key = ""
    if "customer_key" in request.session:
        customer_key = request.session["customer_key"]
  
    logout(request)

    if customer_key != "":
        return redirect("login", customer_key)
    else:
        return redirect("login")

@login_required
def add_instance(request, key):
    info("========In add_instance========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    # If the user is not a staff user or is not a customer user, then show error
    if request.user.is_staff == False:
        if request.user.profile.user_type != "ADMIN":
            return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    if request.method == "POST":
        customer_key = ""
        if "customer_key" in request.session:
            customer_key = request.session["customer_key"]

        customer = Customer.objects.get(customer_key=customer_key)
        instance_name = request.POST.get("instance_name", "")
        instance_type = request.POST.get("instance_type", "")
        instance_url = request.POST.get("instance_url", "")
        instance_user = request.POST.get("instance_user", "")
        instance_password = request.POST.get("instance_password", "")
        #pass_key_enabled = request.POST.get("pass_key_enabled", "")
        pass_key_enabled = "N"
        
        # Create new instance record
        new_instance = Instance(
            customer = customer,
            instance_name = instance_name,
            instance_type = instance_type,
            otm_url = instance_url,
            otm_user = instance_user,
            otm_password = instance_password,
            passkey_enabled = pass_key_enabled
        )

        try:
            new_instance.save()
        except Exception as e:
            if 'unique constraint' in str(e.args).lower():
                return render(request, "rate/common/error_page.html", {"error_message" : f"Instance {instance_name} already exists."})
            else:
                return render(request, "rate/common/error_page.html", {"error_message" : "Error while saving new instance - " + str(e)})

        return redirect("manager-dashboard", key) 
    else:
        return render(request, "rate/customer/superuser/add_instance_page.html")


@login_required
def delete_instance(request, key, id):
    info("========In delete_instance========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    # If the user is not a staff user or is not a customer user, then show error
    if request.user.is_staff == False:
        if request.user.profile.user_type != "ADMIN":
            return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    customer_key = ""
    if "customer_key" in request.session:
        customer_key = request.session["customer_key"]

    instance = Instance.objects.get(id=id)

    if customer_key != instance.customer.customer_key:
        return render(request, "rate/common/error_page.html", {"error_message" : "Unauthorized Access!"})

    instance.delete()  
    
    return redirect("manager-dashboard", key)

@login_required
def update_instance(request, key, id):
    info("========In update_instance========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    # If the user is not a staff user or is not a customer user, then show error
    if request.user.is_staff == False:
        if request.user.profile.user_type != "ADMIN":
            return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    customer_key = ""
    if "customer_key" in request.session:
        customer_key = request.session["customer_key"]

    instance = Instance.objects.get(id=id)

    if customer_key != instance.customer.customer_key:
        return render(request, "rate/common/error_page.html", {"error_message" : "Unauthorized Access!"})
    
    if request.method == "POST":
        instance_name = request.POST.get("instance_name", "")
        instance_type = request.POST.get("instance_type", "")
        instance_url = request.POST.get("instance_url", "")
        instance_user = request.POST.get("instance_user", "")
        instance_password = request.POST.get("instance_password", "")

        instance.instance_name = instance_name
        instance.instance_type = instance_type
        instance.otm_url = instance_url
        instance.otm_user = instance_user
        instance.otm_password = instance_password
        instance.save()
        return redirect("manager-dashboard", key)
    else:
        return render(request, "rate/customer/superuser/edit_instance_page.html", {"instance" : instance})

@login_required
def add_access(request, key):
    info("========In add_access========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    # If the user is not a staff user or is not a customer user, then show error
    if request.user.is_staff == False:
        if request.user.profile.user_type != "ADMIN":
            return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    if request.method == "POST":
        customer_key = ""
        if "customer_key" in request.session:
            customer_key = request.session["customer_key"]

        customer = Customer.objects.get(customer_key=customer_key)

        instance_id = int(request.POST.get("instance_id", ""))
        instance = Instance.objects.get(id=instance_id)
        user_id = int(request.POST.get("user_id", ""))
        user = User.objects.get(id=user_id)
        
        # Create new access record
        new_user_access = UserAccess(
            user = user,
            customer = customer,
            instance = instance
        )

        try:
            new_user_access.save()            
        except IntegrityError as e: 
            if "unique constraint" in str(e.args).lower():
                return render(request, "rate/common/error_page.html", {"error_message" : f"{user.username} is already having access to {instance.instance_name}."})
            else:
                return render(request, "rate/common/error_page.html", {"error_message" : "Error while granting new instance - " + str(e)})
        except Exception as e:
            return render(request, "rate/common/error_page.html", {"error_message" : "Error while granting new instance - " + str(e)})
            
        return redirect("manager-dashboard", key) 

    else:
        customer_key = ""
        if "customer_key" in request.session:
            customer_key = request.session["customer_key"]

        customer = Customer.objects.get(customer_key=customer_key)
        profiles = Profile.objects.filter(customer=customer, user_type="USER").order_by("creation_date")
        if customer.allow_admin_rate_upload == "Y":
            # If admin users are allowed to upload rates for this customer, then select all users
            profiles = Profile.objects.filter(customer=customer).order_by("creation_date")
        instances = Instance.objects.filter(customer=customer).order_by("creation_date")

        return render(request, "rate/customer/superuser/add_access_page.html", {"profiles" : profiles, "instances" : instances})


@login_required
def delete_access(request, key, id):
    info("========In delete_access========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    # If the user is not a staff user or is not a customer user, then show error
    if request.user.is_staff == False:
        if request.user.profile.user_type != "ADMIN":
            return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    customer_key = ""
    if "customer_key" in request.session:
        customer_key = request.session["customer_key"]

    user_access = UserAccess.objects.get(id=id)

    if customer_key != user_access.customer.customer_key:
        return render(request, "rate/common/error_page.html", {"error_message" : "Unauthorized Access!"})

    user_access.delete()  
    
    return redirect("manager-dashboard", key)
  
  
@login_required
def add_profile(request, key):
    info("========In add_profile========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    # If the user is not a staff user or is not a customer user, then show error
    if request.user.is_staff == False:
        if request.user.profile.user_type != "ADMIN":
            return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    if request.method == "POST":
        customer_key = ""
        if "customer_key" in request.session:
            customer_key = request.session["customer_key"]

        customer = Customer.objects.get(customer_key=customer_key)

        user_name = request.POST.get("user_name", "")
        password = request.POST.get("password", "")
        repassword = request.POST.get("repassword", "")
        user_type = request.POST.get("user_type", "")
        group_name = request.POST.get("group_name", "")
        region = request.POST.get("region", "")

        group_names = ""
        if group_name != "":
            group_names = ",".join([(group.strip().upper()) for group in group_name.split(",")]).strip(",")

        if not password == repassword:
            return render(request, "rate/customer/superuser/add_profile_page.html", {"user_name" : user_name, "user_type" : user_type, "group_name" : group_names, "region" : region, "error" : True, "error_message" : "Both passwords must match.", })

        error, error_message = validate_password_stregnth(password)
        if error:
            return render(request, "rate/customer/superuser/add_profile_page.html", {"user_name" : user_name, "user_type" : user_type, "group_name" : group_names, "region" : region, "error" : True, "error_message" : error_message })
        
        # Create new user
        try:
            new_user = User.objects.create_user(
                username = user_name,
                email = user_name,
                password = password
            )
        except IntegrityError as e: 
            if "unique constraint" in str(e.args).lower():
                return render(request, "rate/customer/superuser/add_profile_page.html", {"user_name" : user_name, "user_type" : user_type, "group_name" : group_names, "error" : True, "error_message" : f"User Name {user_name} already exists." })
            else:
                return render(request, "rate/common/error_page.html", {"error_message" : "Error while creating new user - " + str(e)})
        except Exception as e:
            return render(request, "rate/common/error_page.html", {"error_message" : "Error while creating new user - " + str(e)})        

        # If no group name was provided for new user, make it DEFAULT
        if group_names == "":
            group_names = "DEFAULT"

        # Create profile record
        new_profile = Profile(
            user = new_user,
            customer = customer,
            user_type = user_type,
            group_name = group_names,
            region = region
        )

        try:
            new_user.save()
            new_profile.save()

            # Split the group names
            for group_name in group_names.upper().split(","):
                if group_name.strip() != "":
                    groups = Group.objects.filter(customer=customer, group_name=group_name)
                    # If the group name is not already defined for the customer, then add it
                    if len(groups) == 0:
                        group = Group(customer=customer, group_name=group_name)
                        group.save()

        except IntegrityError as e: 
            return render(request, "rate/common/error_page.html", {"error_message" : "Error while creating new profile - " + str(e)})
        except Exception as e:
            return render(request, "rate/common/error_page.html", {"error_message" : "Error while creating new profile - " + str(e)})

        return redirect("manager-dashboard", key) 
    else:
        return render(request, "rate/customer/superuser/add_profile_page.html", {"user_type" : "USER", "group_name" : "DEFAULT"})

@login_required
def delete_profile(request, key, id):
    info("========In delete_profile========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    # If the user is not a staff user or is not a customer user, then show error
    if request.user.is_staff == False:
        if request.user.profile.user_type != "ADMIN":
            return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    customer_key = ""
    if "customer_key" in request.session:
        customer_key = request.session["customer_key"]

    profile = Profile.objects.get(id=id)
    user = profile.user

    if customer_key != profile.customer.customer_key:
        return render(request, "rate/common/error_page.html", {"error_message" : "Unauthorized Access!"})

    user.delete()  
    
    return redirect("manager-dashboard", key)

@login_required
def update_profile(request, key, id):
    info("========In update_profile========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    # If the user is not a staff user or is not a customer user, then show error
    if request.user.is_staff == False:
        if request.user.profile.user_type != "ADMIN":
            return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    customer_key = ""
    if "customer_key" in request.session:
        customer_key = request.session["customer_key"]

    profile = Profile.objects.get(id=id)

    if customer_key != profile.customer.customer_key:
        return render(request, "rate/common/error_page.html", {"error_message" : "Unauthorized Access!"})

    if request.method == "POST":
        group_name = request.POST.get("group_name", "")
        region = request.POST.get("region", "")

        group_names = ",".join([(group.strip().upper()) for group in group_name.split(",")]).strip(",")

        # If no group name was provided for new user, make it DEFAULT
        if group_names == "":
            group_names = "DEFAULT"
        
        profile.group_name = group_names.strip(",")
        profile.region = region        

        try:
            profile.save()
        except Exception as e:
            return render(request, "rate/common/error_page.html", {"error_message" : "Error while creating new profile - " + str(e)})
        
        # If the group name is not already defined for the customer, then add it
        for group_name in group_names.split(","):
            if group_name.strip() != "":
                groups = Group.objects.filter(customer=profile.customer, group_name=group_name)
                # If the group name is not already defined for the customer, then add it
                if len(groups) == 0:
                    group = Group(customer=profile.customer, group_name=group_name)
                    group.save()
        
        return redirect("manager-dashboard", key) 
    else:
        return render(request, "rate/customer/superuser/edit_profile_page.html", {"profile" : profile})
    
@login_required
def reset_password(request, key, id):
    info("========In reset_password========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    # If the user is not a staff user or is not a customer user, then show error
    if request.user.is_staff == False:
        if request.user.profile.user_type != "ADMIN":
            return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    customer_key = ""
    if "customer_key" in request.session:
        customer_key = request.session["customer_key"]

    # Get the profile for whom password should be reset
    profile = Profile.objects.get(id=id)
    current_user = profile.user # current_user is that user whose password should be changed

    if customer_key != profile.customer.customer_key:
        return render(request, "rate/common/error_page.html", {"error_message" : "Unauthorized Access!"})

    if request.method == "POST":
        password = request.POST.get("password", "")
        repassword = request.POST.get("repassword", "")

        if not password == repassword:
            return render(request, "rate/customer/superuser/reset_password_page.html", {"user_name" : current_user.username, "user_type" : profile.user_type, "error" : True, "error_message" : "Both passwords must match." })

        error, error_message = validate_password_stregnth(password)
        if error:
            return render(request, "rate/customer/superuser/reset_password_page.html", {"user_name" : current_user.username, "user_type" : profile.user_type, "error" : error, "error_message" : error_message })
        
        try:
            current_user.set_password(password)
            current_user.save()
        except Exception as e:
            return render(request, "rate/common/error_page.html", {"error_message" : "Error while resetting password - " + str(e)})
        
        return redirect("manager-dashboard", key) 
    else:
        return render(request, "rate/customer/superuser/reset_password_page.html", {"user" : current_user, "user_name" : current_user.username})

@login_required
def change_password(request, key):
    info("========In change_password========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    customer_key = ""
    if "customer_key" in request.session:
        customer_key = request.session["customer_key"]

    profile = request.user.profile
    current_user = request.user

    if customer_key != profile.customer.customer_key:
        return render(request, "rate/common/error_page.html", {"error_message" : "Unauthorized Access!"})

    if request.method == "POST":
        current_password = request.POST.get("current_password", "")
        password = request.POST.get("password", "")
        repassword = request.POST.get("repassword", "")

        if not current_user.check_password(current_password):
            return render(request, "rate/common/change_password_page.html", {"user_name" : current_user.username, "user_type" : profile.user_type, "error" : True, "error_message" : "Current password is incorrect." })

        if not password == repassword:
            return render(request, "rate/common/change_password_page.html", {"user_name" : current_user.username, "user_type" : profile.user_type, "error" : True, "error_message" : "Both passwords must match." })

        error, error_message = validate_password_stregnth(password)
        if error:
            return render(request, "rate/common/change_password_page.html", {"user_name" : current_user.username, "user_type" : profile.user_type, "error" : error, "error_message" : error_message })
        
        try:
            current_user.set_password(password)
            current_user.save()
            return redirect("login", customer_key)
        except Exception as e:
            return render(request, "rate/common/error_page.html", {"error_message" : "Error while changing password - " + str(e)})
        
        return redirect("manager-dashboard", key) 
    else:
        return render(request, "rate/common/change_password_page.html", {"user" : current_user, "user_name" : current_user.username})  

@login_required
def master_admin_customers(request):
    info("========In master_admin_customers========")
    info("Logged in user name " + request.user.username)
    # If the super user has logged in then the home page shouldn't be displayed
    # because the super user is not linked to any customer
    if request.user.is_staff == True:
        if request.method == 'POST':
            if "reset_data" in request.POST:
                info("Resetting data...")
                customer_key = request.POST.get("customer", "")
                customer = Customer.objects.get(customer_key=customer_key)
                master_customer = Customer.objects.get(domain="MASTER")
                message = ""
                outcome = ""
                try:
                    Template.objects.filter(customer=customer).delete()
                    TemplateStaticColumn.objects.filter(customer=customer).delete()
                    TemplateDynamicColumnValue.objects.filter(customer=customer).delete()
                    TemplateDynamicColumn.objects.filter(customer=customer).delete()                
                    Batch.objects.filter(customer=customer).delete()
                    for csv_file in CsvFile.objects.filter(customer=customer):
                        CsvStructure.objects.filter(csv_file=csv_file).delete()
                    CsvFile.objects.filter(customer=customer).delete()
                    Group.objects.filter(customer=customer).delete()

                    cursor = connection.cursor()

                    # Delete all customer specific records from all relevant tables
                    cursor.execute("UPDATE sqlite_sequence SET seq = (SELECT MAX(id) FROM rate_template) WHERE LOWER(name) = 'rate_template'")
                    cursor.execute("UPDATE sqlite_sequence SET seq = (SELECT MAX(id) FROM rate_templatestaticcolumn) WHERE LOWER(name) = 'rate_templatestaticcolumn'")
                    cursor.execute("UPDATE sqlite_sequence SET seq = (SELECT MAX(id) FROM rate_templatedynamiccolumn) WHERE LOWER(name) = 'rate_templatedynamiccolumn'")
                    cursor.execute("UPDATE sqlite_sequence SET seq = (SELECT MAX(id) FROM rate_templatedynamiccolumnvalue) WHERE LOWER(name) = 'rate_templatedynamiccolumnvalue'")
                    cursor.execute("UPDATE sqlite_sequence SET seq = (SELECT MAX(id) FROM rate_csvfile) WHERE LOWER(name) = 'rate_csvfile'")
                    cursor.execute("UPDATE sqlite_sequence SET seq = (SELECT MAX(id) FROM rate_csvstructure) WHERE LOWER(name) = 'rate_csvstructure'")
                    cursor.execute("UPDATE sqlite_sequence SET seq = (SELECT MAX(id) FROM rate_group) WHERE LOWER(name) = 'rate_group'")
                    transaction.commit()

                    # Create DEFAULT group
                    group = Group(customer=customer, group_name="DEFAULT") # Add default group
                    group.save()

                    # Get master templates which should be copied
                    master_templates = Template.objects.filter(customer=master_customer)
                    for master_template in master_templates:
                        # Create new template
                        template = Template(
                            customer=customer, 
                            offering_type_name=master_template.offering_type_name,
                            offering_type_description=master_template.offering_type_description,
                            rate_type_name=master_template.rate_type_name,
                            group_name="DEFAULT",
                            enabled=master_template.enabled
                        )
                        template.save()
                        
                        # Get static columns for master customer
                        master_static_columns = TemplateStaticColumn.objects.filter(customer=master_customer, template=master_template)
                        for master_static_column in master_static_columns:
                            # Create new static column
                            staticColumn = TemplateStaticColumn(
                                customer=customer, 
                                template=template, 
                                template_column_name=master_static_column.template_column_name, 
                                master_column_name=master_static_column.master_column_name, 
                                enterable=master_static_column.enterable, 
                                required=master_static_column.required, 
                                default_value=master_static_column.default_value, 
                                enabled=master_static_column.enabled, 
                                include_in_template=master_static_column.include_in_template, 
                                position=master_static_column.position
                            )
                            staticColumn.save()

                        # Get dynamic columns for master customer
                        master_dyanmic_columns = TemplateDynamicColumn.objects.filter(customer=master_customer, template=master_template)
                        for master_dyanmic_column in master_dyanmic_columns:
                            # Create new dynamic column
                            dynamicColumn = TemplateDynamicColumn(
                                customer=customer, 
                                template=template, 
                                parameter_name=master_dyanmic_column.parameter_name, 
                                enabled=master_dyanmic_column.enabled, 
                                position=master_dyanmic_column.position
                            )
                            dynamicColumn.save()

                            # Get dynamic column values for master customer
                            master_dyanmic_column_values = TemplateDynamicColumnValue.objects.filter(customer=master_customer, template=master_template, param=master_dyanmic_column)
                            for master_dyanmic_column_value in master_dyanmic_column_values:
                                # Create new dynamic column value
                                dynamic_column_value = TemplateDynamicColumnValue(
                                    customer=customer, 
                                    template=template, 
                                    param=dynamicColumn, 
                                    sequence=master_dyanmic_column_value.sequence, 
                                    parameter_value=master_dyanmic_column_value.parameter_value, 
                                    default_selected=master_dyanmic_column_value.default_selected, 
                                    template_column_names=master_dyanmic_column_value.template_column_names, 
                                    master_column_names=master_dyanmic_column_value.master_column_names, 
                                    enabled=master_dyanmic_column_value.enabled
                                )
                                dynamic_column_value.save()

                    # Get csv files for master customer
                    master_csv_files = CsvFile.objects.filter(customer=master_customer)
                    for master_csv_file in master_csv_files:
                        # Create new csv file
                        csv_file = CsvFile(
                            customer=customer, 
                            sequence=master_csv_file.sequence, 
                            name=master_csv_file.name, 
                            csv_file_name=master_csv_file.csv_file_name, 
                            description=master_csv_file.description, 
                            file_identifier=master_csv_file.file_identifier, 
                            enabled=master_csv_file.enabled, 
                            auto_remove_duplicates=master_csv_file.auto_remove_duplicates, 
                            auto_number_initial_sequence=master_csv_file.auto_number_initial_sequence, 
                            has_date_field=master_csv_file.has_date_field,
                            date_format=master_csv_file.date_format,
                            csv_command=master_csv_file.csv_command
                        )
                        csv_file.save()

                        # Get csv file fields for master customer
                        master_csv_structures = CsvStructure.objects.filter(csv_file=master_csv_file)
                        for master_csv_structure in master_csv_structures:
                            # Create new csv file field
                            csv_structure = CsvStructure(
                                csv_file=csv_file, 
                                sequence=master_csv_structure.sequence, 
                                column_name=master_csv_structure.column_name, 
                                source_column_name=master_csv_structure.source_column_name, 
                                fixed_value=master_csv_structure.fixed_value, 
                                expression=master_csv_structure.expression, 
                                enabled=master_csv_structure.enabled
                            )
                            csv_structure.save()

                    message = "All configurations have been reset for customer " + customer.customer_name + "."
                    outcome = "SUCCESS"
                except Exception as ex:
                    message = "Error encountered while resetting all configurations for customer " + customer.customer_name + " - " + str(ex)
                    outcome = "ERROR"

                info("Outcome " + message)
                customers = Customer.objects.filter(~Q(domain="MASTER"))

                full_url = request.POST.get("system_url", "")
                full_url = full_url.replace(reverse("master-customers"), "/") # This will give the base url of the application, will be used to display customer specific urls in the page
                templates = Template.objects.filter(Q(customer__in=customers))
                return render(request, "rate/admin/admin_view_customer.html", {"customers" : customers, "full_url" : full_url, "templates" : templates, "reset_message" : message, "reset_outcome" : outcome})
            
            elif "clone_data" in request.POST:
                info("Cloning data...")
                template_id = int(request.POST.get("template_id", ""))
                group_name = request.POST.get("group", "")
                group_name = group_name.upper()

                master_template = Template.objects.get(id=template_id)
                customer = master_template.customer

                # If the provided group name doesn't already exists for the customer, then add it
                group = Group.objects.none()
                groups = Group.objects.filter(customer=customer, group_name=group_name)
                if len(groups) == 0:
                    group = Group(customer=customer, group_name=group_name)
                    group.save()
                else:
                    group = groups[0]

                try:
                    # Create new template
                    template = Template(
                                customer=customer, 
                                offering_type_name=master_template.offering_type_name,
                                offering_type_description=master_template.offering_type_description,
                                rate_type_name=master_template.rate_type_name,
                                group_name=group_name,
                                enabled=master_template.enabled
                            )
                    template.save()

                    # Get static columns from the source static column
                    master_static_columns = TemplateStaticColumn.objects.filter(customer=customer, template=master_template)
                    for master_static_column in master_static_columns:
                        # Create static column
                        staticColumn = TemplateStaticColumn(
                            customer=customer, 
                            template=template, 
                            template_column_name=master_static_column.template_column_name, 
                            master_column_name=master_static_column.master_column_name, 
                            enterable=master_static_column.enterable, 
                            required=master_static_column.required, 
                            default_value=master_static_column.default_value, 
                            enabled=master_static_column.enabled, 
                            include_in_template=master_static_column.include_in_template, 
                            position=master_static_column.position
                        )
                        staticColumn.save()

                    # Get dynamic columns from the source static column
                    master_dyanmic_columns = TemplateDynamicColumn.objects.filter(customer=customer, template=master_template)
                    for master_dyanmic_column in master_dyanmic_columns:
                        # Create dynamic column
                        dynamicColumn = TemplateDynamicColumn(
                            customer=customer, 
                            template=template, 
                            parameter_name=master_dyanmic_column.parameter_name, 
                            enabled=master_dyanmic_column.enabled, 
                            position=master_dyanmic_column.position
                        )
                        dynamicColumn.save()

                        # Get dynamic column values from the source static column
                        master_dyanmic_column_values = TemplateDynamicColumnValue.objects.filter(customer=customer, template=master_template, param=master_dyanmic_column)
                        for master_dyanmic_column_value in master_dyanmic_column_values:
                            # Create dynamic column value
                            dynamic_column_value = TemplateDynamicColumnValue(
                                customer=customer, 
                                template=template, 
                                param=dynamicColumn, 
                                sequence=master_dyanmic_column_value.sequence, 
                                parameter_value=master_dyanmic_column_value.parameter_value, 
                                default_selected=master_dyanmic_column_value.default_selected, 
                                template_column_names=master_dyanmic_column_value.template_column_names, 
                                master_column_names=master_dyanmic_column_value.master_column_names, 
                                enabled=master_dyanmic_column_value.enabled
                            )
                            dynamic_column_value.save()   

                    #message = f"Cloning is complete and new template has been created for {customer.customer_name} - {template.offering_type_name} - {template.rate_type_name} - {template.group_name}."
                    message = f"Template created for {customer.customer_name} - {template.offering_type_name} - {template.rate_type_name} - {template.group_name}."
                    outcome = "SUCCESS"
                    
                except Exception as ex:
                    if 'unique constraint' in str(ex.args).lower():
                        message = "Same template already exists for the entered group name."
                        outcome = "ERROR"
                    else:
                        message = "Same template already exists for the entered group name."
                        outcome = "ERROR"

                info("Outcome " + message)
                customers = Customer.objects.filter(~Q(domain="MASTER"))

                full_url = request.POST.get("system_url", "")
                print(full_url)
                full_url = full_url.replace(reverse("master-customers"), "/")
                templates = Template.objects.filter(Q(customer__in=customers))
                return render(request, "rate/admin/admin_view_customer.html", {"customers" : customers, "full_url" : full_url, "templates" : templates, "clone_message" : message, "clone_outcome" : outcome})  
        else:
            customers = Customer.objects.filter(~Q(domain="MASTER"))
            full_url = request.POST.get("system_url", "")
            if full_url == "":
                full_url = request.build_absolute_uri()
                full_url = full_url.replace(reverse("master-customers"), "/")

            templates = Template.objects.filter(Q(customer__in=customers))
            return render(request, "rate/admin/admin_view_customer.html", {"customers" : customers, "full_url" : full_url, "templates" : templates})
    else:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

def master_admin_manage_customer(request, key):
    info("========In master_admin_manage_customer========")

    # If staff user then redirect to admin customer page
    if request.user.is_staff:
        info("Logged in user is a staff user")
        request.session["customer_key"] = key
        
        # Get the customer
        customer = Customer.objects.get(customer_key=key)
        # Store customer key, customer name and domain from user profile, and store them in session
        info("Customer name " + customer.customer_name)
        request.session["customer_key"] = customer.customer_key
        request.session["customer_name"] = customer.customer_name
        request.session["domain"] = customer.domain
        request.session["allow_admin_rate_upload"] = "Y"

        # Navigate to manager dashboard of the customer
        return redirect("manager-dashboard", key)
    else:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})
    
    return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})


def master_admin_edit_customer(request, key):
    info("========In master_admin_edit_customer========")

    # If staff user then redirect to admin customer page
    if not request.user.is_staff:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    info("Logged in user is a staff user")
    # Get the customer
    customer = Customer.objects.get(customer_key=key)

    if request.method == "POST":
        customer_name = request.POST.get("customer_name", "")
        domain = request.POST.get("domain", "")
        end_date = request.POST.get("end_date", "")
        allow_admin_rate_upload = request.POST.get("allow_admin_rate_upload", "N")
        view_all_template_batches = request.POST.get("view_all_template_batches", "N")

        # Update customer record
        customer.customer_name = customer_name
        customer.domain = domain
        customer.allow_admin_rate_upload = allow_admin_rate_upload
        customer.view_all_template_batches = view_all_template_batches
        if end_date:
            customer.end_date = end_date

        customer.save()
        return redirect("master-customers")
    else:
        # Changing format of the date so that it can be defaulted in the html
        end_date = customer.end_date
        if end_date:
            end_date = end_date.strftime("%Y-%m-%d")
        return render(request, "rate/admin/admin_edit_customer.html", {"customer" : customer, "end_date" : end_date})
    
    return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})


def master_admin_add_customer(request):
    info("========In master_admin_add_customer========")

    # If staff user then redirect to admin customer page
    if not request.user.is_staff:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    info("Logged in user is a staff user")
   
    if request.method == "POST":
        customer_name = request.POST.get("customer_name", "")
        domain = request.POST.get("domain", "")
        end_date = request.POST.get("end_date", "")
        allow_admin_rate_upload = request.POST.get("allow_admin_rate_upload", "N")
        view_all_template_batches = request.POST.get("view_all_template_batches", "N")

        # Add customer record
        customer = Customer.objects.none()

        try:
            if end_date:
                customer = Customer(customer_name=customer_name, domain=domain, allow_admin_rate_upload=allow_admin_rate_upload, view_all_template_batches=view_all_template_batches, end_date=end_date)
            else:
                customer = Customer(customer_name=customer_name, domain=domain, allow_admin_rate_upload=allow_admin_rate_upload, view_all_template_batches=view_all_template_batches)
        except Exception as ex:
            if 'unique constraint' in str(ex.args).lower():
                return render(request, "rate/admin/admin_add_customer.html", {"error_message" : "Customer with same name already exists."})
            else:
                return render(request, "rate/common/error_page.html", {"error_message" : "Error while adding new customer - " + str(ex)})
        
        customer.save()
        return redirect("master-customers")
    else:
        return render(request, "rate/admin/admin_add_customer.html", {})
    
    return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})


def master_admin_delete_template(request, key, id):
    info("========In master_admin_delete_template========")

    # If staff user then redirect to admin customer page
    if not request.user.is_staff:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    info("Logged in user is a staff user")

    try:
        template = Template.objects.get(id=id)
        template.delete()
    except Exception as ex:
        return render(request, "rate/common/error_page.html", {"error_message" : "Error while deleting template - " + str(ex)})
    
    return redirect("manager-dashboard", key)
    

def get_contracts_ajax(request):
    info("========In get_contracts_ajax========")
    if request.method == "POST":
        try:
            contracts = []
            json_data = json.loads(request.body)
            service_provider = json_data.get("service_provider")
            info("Service Provider in AJAX = " + service_provider)

            if not service_provider or service_provider == "":
                return contracts

            customer_key = ""
                
            # Get customer key from the session and fetch customer object
            if "customer_key" in request.session:
                customer_key = request.session["customer_key"]

            customer = Customer.objects.get(customer_key=customer_key)            

            # Get instance details
            instance_id = 0
            if "instance_id" in request.session:
                instance_id = int(request.session["instance_id"])            

            instance = Instance.objects.get(id=instance_id, customer=customer)

            # Get mode from session
            offering_type_name = ""
            offering_type_description = ""
            if "offering_type_name" in request.session:
                offering_type_name = request.session["offering_type_name"]
            if "offering_type_description" in request.session:
                offering_type_description = request.session["offering_type_description"]
            mode = offering_type_description if offering_type_description != "" else offering_type_name

            # Get list of contracts          
            for contract in get_contracts(instance, service_provider, customer.domain, mode):
                contracts.append({"id" : contract, "value" : contract})

            info("Returning " + str(contracts))

            return JsonResponse(contracts, safe=False)

        except Exception as e:
            error("get_contracts_ajax - " + str(e))
            data = {}
            data["error_message"] = "error"
            return JsonResponse(data)

        return JsonResponse(contracts, safe = False) 

def get_rate_types_ajax(request):
    info("========In get_rate_types_ajax========")
    if request.method == "POST":
        try:
            batches = []
            json_data = json.loads(request.body)
            offering_type_group_name = json_data.get("offering_type_group_name")
            #template_id = request.POST["template_id"]
            offering_type=offering_type_group_name.split("-")[0]
            group_name=offering_type_group_name.split("-")[1]
            info("Offering Type in AJAX " + offering_type)
            info("Group Name in AJAX " + group_name)

            customer_key = ""                
            # Get customer key from the session and fetch customer object
            if "customer_key" in request.session:
                customer_key = request.session["customer_key"]

            customer = Customer.objects.get(customer_key=customer_key) 
            templates = Template.objects.filter(customer=customer, offering_type_name=offering_type, group_name=group_name).order_by("rate_type_name")

            # Get list of batches
            offering_types = []
            for template in templates:
                offering_types.append({"id" : template.id, "value" : template.rate_type_name})

            info("Returning " + str(offering_types))
            return JsonResponse(offering_types, safe=False)

        except Exception as e:
            error("get_rate_types_ajax - " + str(e))
            data = {}
            data["error_message"] = "error"
            return JsonResponse(data)

        return JsonResponse([], safe = False)

# def get_template_batches_ajax(request):
#     info("========In get_template_batches_ajax========")
#     if request.method == "POST":
#         try:
#             batches = []
#             json_data = json.loads(request.body)
#             template_id = json_data.get("template_id")
#             #template_id = request.POST["template_id"]
#             info("Template Id in AJAX " + template_id)

#             customer_key = ""                
#             # Get customer key from the session and fetch customer object
#             if "customer_key" in request.session:
#                 customer_key = request.session["customer_key"]

#             customer = Customer.objects.get(customer_key=customer_key) 
#             template = Template.objects.get(id=int(template_id))

#             #batches = Batch.objects.filter(customer=customer, template=template).order_by('-id')
#             batches = Batch.objects.filter(customer=customer).order_by('-id')

#             # Get list of batches
#             template_batches = []
#             for batch in batches:
#                 template_batches.append({"id" : str(batch.id), "value" : batch.batch_name})

#             info("Returning " + str(template_batches))
#             return JsonResponse(template_batches, safe=False)

#         except Exception as e:
#             error("get_template_batches_ajax - " + str(e))
#             data = {}
#             data["error_message"] = "error"
#             return JsonResponse(data)

#         return JsonResponse([], safe = False)


@login_required
def delete_batch(request, key, id):
    info("========In delete_batch========")
    if not request.user.is_authenticated: 
        return HttpResponseRedirect("login")

    # If customer key from session doesn't match with the key in the URL, then show error
    if not key == request.session["customer_key"]:
        return render(request, "rate/common/error_page.html", {"error_message" : "You are not authorized to view this page."})

    customer_key = ""
    if "customer_key" in request.session:
        customer_key = request.session["customer_key"]

    batch = Batch.objects.get(id=id)

    batch.delete()  
    
    return redirect("manager-dashboard", key)


def page_not_found(request, exception):
    info("========In page_not_found========")
    return render(request, 'rate/common/404.html', {})